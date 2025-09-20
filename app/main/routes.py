import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from flask import (render_template, request, jsonify, redirect, 
                   url_for, flash, Response, make_response)
from flask_login import login_required, current_user
from datetime import date, timedelta, datetime
from sqlalchemy import and_, func, case
import openpyxl
import json
import bleach

from . import main_bp
from app import db
from app.models import WorkSchedule, ActivityLog, Personnel, TaskAssignment
from app.utils import log_activity
from app.decorators import permission_required, admin_required

# --- HTML 清洗配置 ---
ALLOWED_TAGS = ['p', 'b', 'strong', 'i', 'em', 'span', 'br']
ALLOWED_ATTRIBUTES = {'span': ['style']}
ALLOWED_STYLES = ['color']

def sanitize_html(html_content):
    """使用 Bleach 清洗用户输入的 HTML，防止 XSS 攻击。"""
    if not html_content:
        return ""
    return bleach.clean(
        html_content,
        tags=ALLOWED_TAGS,
        attributes=ALLOWED_ATTRIBUTES,
        styles=ALLOWED_STYLES,
        strip=True
    ).strip()
# --------------------


def get_week_dates(start_date_str=None):
    if start_date_str:
        try:
            base_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            base_date = date.today()
    else:
        base_date = date.today()
    
    start_of_week = base_date - timedelta(days=base_date.weekday())
    return [start_of_week + timedelta(days=i) for i in range(7)]

@main_bp.route('/')
@login_required
def index():
    start_date_str = request.args.get('start_date')
    week_dates = get_week_dates(start_date_str)
    start_of_week = week_dates[0]
    end_of_week = week_dates[-1]

    today = date.today()
    current_week_start = today - timedelta(days=today.weekday())
    is_current_week = (start_of_week == current_week_start)

    tasks = WorkSchedule.query.filter(
        WorkSchedule.is_deleted == False, # 过滤掉已删除的任务
        WorkSchedule.task_date.between(start_of_week, end_of_week)
    ).order_by(WorkSchedule.task_date, WorkSchedule.position).all()
    
    personnel_list = Personnel.query.order_by(Personnel.name).all()
    
    schedule_by_day = {day: [] for day in week_dates}
    for task in tasks:
        if task.task_date in schedule_by_day:
            schedule_by_day[task.task_date].append(task)
        
    # 判断周末是否有任务，用于前端提示
    saturday_date = week_dates[5]
    sunday_date = week_dates[6]
    weekend_has_tasks = len(schedule_by_day[saturday_date]) > 0 or len(schedule_by_day[sunday_date]) > 0
        
    prev_week_start = start_of_week - timedelta(days=7)
    next_week_start = start_of_week + timedelta(days=7)

    page_main_title = "周工作计划"
    page_date_range = f"({start_of_week.strftime('%Y-%m-%d')} 至 {end_of_week.strftime('%Y-%m-%d')})"

    return render_template(
        'main/index.html', 
        schedule_by_day=schedule_by_day,
        week_dates=week_dates,
        personnel_list=personnel_list,
        prev_week=prev_week_start.strftime('%Y-%m-%d'),
        next_week=next_week_start.strftime('%Y-%m-%d'),
        is_current_week=is_current_week,
        page_main_title=page_main_title,
        page_date_range=page_date_range,
        weekend_has_tasks=weekend_has_tasks
    )

@main_bp.route('/add_task', methods=['POST'])
@login_required
@permission_required('can_add')
def add_task():
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date') or start_date_str
    content = request.form.get('content')
    personnel_json_str = request.form.get('personnel', '[]')
    
    sanitized_content = sanitize_html(content)

    try:
        personnel_list = json.loads(personnel_json_str)
        personnel_names = [item['value'] for item in personnel_list if item.get('value')]
    except (json.JSONDecodeError, TypeError):
        personnel_names = []

    if not all([start_date_str, sanitized_content, personnel_names]):
        flash('开始日期、工作内容和人员均为必填项。', 'danger')
        return redirect(url_for('main.index'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('日期格式不正确。', 'danger')
        return redirect(url_for('main.index', start_date=start_date_str))

    current_date = start_date
    while current_date <= end_date:
        max_pos = db.session.query(func.max(WorkSchedule.position)).filter_by(task_date=current_date, is_deleted=False).scalar() or -1
        new_task = WorkSchedule(
            task_date=current_date,
            content=sanitized_content,
            author_id=current_user.id,
            position=max_pos + 1
        )
        for index, name in enumerate(personnel_names):
            assignment = TaskAssignment(personnel_name=name, position=index)
            new_task.assignments.append(assignment)
        db.session.add(new_task)
        current_date += timedelta(days=1)
    
    db.session.commit()
    log_activity('创建任务', f"为日期 {start_date_str} 到 {end_date_str} 添加了任务: '{sanitized_content}'")
    flash('新任务已添加。', 'success')
        
    return redirect(url_for('main.index', start_date=start_date_str))

@main_bp.route('/api/get_task_with_range/<int:task_id>')
@login_required
def get_task_with_range(task_id):
    task = WorkSchedule.query.filter_by(id=task_id, is_deleted=False).first_or_404()
    personnel_set = set(a.personnel_name for a in task.assignments)
    
    start_date = task.task_date
    current_date = task.task_date - timedelta(days=1)
    while True:
        prev_task = WorkSchedule.query.filter_by(task_date=current_date, content=task.content, is_deleted=False).first()
        if prev_task and set(a.personnel_name for a in prev_task.assignments) == personnel_set:
            start_date = current_date
            current_date -= timedelta(days=1)
        else:
            break
            
    end_date = task.task_date
    current_date = task.task_date + timedelta(days=1)
    while True:
        next_task = WorkSchedule.query.filter_by(task_date=current_date, content=task.content, is_deleted=False).first()
        if next_task and set(a.personnel_name for a in next_task.assignments) == personnel_set:
            end_date = current_date
            current_date += timedelta(days=1)
        else:
            break
            
    return jsonify({
        'id': task.id,
        'content': task.content,
        'personnel': [a.personnel_name for a in task.assignments],
        'version': task.version,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d')
    })

@main_bp.route('/api/update_task/<int:task_id>', methods=['POST'])
@login_required
@permission_required('can_edit')
def update_task(task_id):
    task = WorkSchedule.query.get_or_404(task_id)
    data = request.get_json()

    if not data:
        return jsonify({'success': False, 'error': '无效数据'}), 400

    client_version = data.get('version')
    if client_version is not None and task.version != client_version:
        current_personnel_list = [a.personnel_name for a in task.assignments]
        return jsonify({
            'success': False, 'error': '此任务已被他人修改，请解决冲突。','conflict': True,
            'current_data': {
                'content': task.content,
                'personnel': current_personnel_list,
                'version': task.version
            }
        }), 409

    raw_content = data.get('content', task.content)
    task.content = sanitize_html(raw_content)
    
    personnel_data = data.get('personnel', None)
    if personnel_data is not None and isinstance(personnel_data, list):
        task.assignments.clear()
        personnel_names = [item['value'] for item in personnel_data if isinstance(item, dict) and item.get('value')]
        for index, name in enumerate(personnel_names):
            task.assignments.append(TaskAssignment(personnel_name=name, position=index))
    
    new_date_str = data.get('task_date')
    if new_date_str:
        try:
            task.task_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': '无效的日期格式'}), 400

    task.version += 1
    db.session.commit()
    log_activity('更新任务', f"更新了任务ID {task_id}，内容: '{task.content}'")
    return jsonify({'success': True, 'message': '任务已更新。'})

@main_bp.route('/api/reorder_tasks', methods=['POST'])
@login_required
@permission_required('can_edit')
def reorder_tasks():
    data = request.get_json()
    moved_task_data = data.get('moved_task')
    target_list_data = data.get('target_list')

    if not moved_task_data or not target_list_data:
        return jsonify({'success': False, 'error': '请求数据不完整。'}), 400

    moved_task_id = moved_task_data['id']
    client_version = moved_task_data['version']

    moved_task_db = WorkSchedule.query.get(moved_task_id)
    if not moved_task_db or moved_task_db.version != client_version:
        return jsonify({'success': False, 'error': '操作失败，任务已被他人修改。请刷新页面后重试。', 'conflict': True}), 409

    try:
        new_date_str = target_list_data.get('date')
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        if moved_task_db.task_date != new_date:
            moved_task_db.task_date = new_date

        all_task_ids = [t['id'] for t in data.get('all_tasks_in_view', [])]
        db.session.query(WorkSchedule).filter(WorkSchedule.id.in_(all_task_ids)).update(
            {'version': WorkSchedule.version + 1}, synchronize_session=False
        )

        target_ids = target_list_data.get('task_ids', [])
        if target_ids:
            target_case = case({int(tid): i for i, tid in enumerate(target_ids)}, value=WorkSchedule.id)
            db.session.query(WorkSchedule).filter(WorkSchedule.id.in_(target_ids)).update(
                {'position': target_case}, synchronize_session=False
            )
        
        db.session.commit()
        log_activity('拖拽任务', f"移动了任务ID {moved_task_id}")
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        log_activity('拖拽任务失败', f"任务ID {moved_task_id} 移动失败: {str(e)}")
        return jsonify({'success': False, 'error': '数据库操作失败。'}), 500

@main_bp.route('/api/delete_task/<int:task_id>', methods=['POST'])
@login_required
@permission_required('can_delete')
def api_delete_task(task_id):
    task = WorkSchedule.query.get_or_404(task_id)
    task.is_deleted = True
    task.deleted_at = datetime.utcnow()
    db.session.commit()
    log_activity('软删除任务', f"软删除了任务ID {task_id}, 内容: '{task.content}'")
    return jsonify({'success': True, 'message': '任务已删除。', 'task_id': task.id})

@main_bp.route('/api/restore_task/<int:task_id>', methods=['POST'])
@login_required
@permission_required('can_add')
def restore_task(task_id):
    task = WorkSchedule.query.get_or_404(task_id)
    if task.is_deleted:
        task.is_deleted = False
        task.deleted_at = None
        db.session.commit()
        log_activity('恢复任务', f"恢复了任务ID {task_id}, 内容: '{task.content}'")
        return jsonify({'success': True, 'message': '任务已恢复。'})
    return jsonify({'success': False, 'error': '任务未被删除。'}), 400

@main_bp.route('/export_excel', methods=['POST'])
@login_required
def export_excel():
    start_date_str = request.form.get('start_date')
    week_dates = get_week_dates(start_date_str)
    start_of_week = week_dates[0]
    end_of_week = week_dates[-1]

    tasks = WorkSchedule.query.filter(
        WorkSchedule.is_deleted == False,
        WorkSchedule.task_date.between(start_of_week, end_of_week)
    ).order_by(WorkSchedule.task_date, WorkSchedule.position).all()
    
    if not tasks:
        flash('该周没有可导出的数据。', 'warning')
        return redirect(url_for('main.index', start_date=start_date_str))

    schedule_by_day = {day: [] for day in week_dates}
    for task in tasks:
        if task.task_date in schedule_by_day:
            personnel_str = ", ".join([a.personnel_name for a in task.assignments])
            schedule_by_day[task.task_date].append(task.content)
            schedule_by_day[task.task_date].append(personnel_str)

    max_rows = max(len(day_tasks) for day_tasks in schedule_by_day.values()) if any(schedule_by_day.values()) else 0
    if max_rows == 0:
        flash('该周没有可导出的数据。', 'warning')
        return redirect(url_for('main.index', start_date=start_date_str))

    data_dict = {}
    for day in week_dates:
        day_str = day.strftime('%Y-%m-%d') + " (星期" + ['一','二','三','四','五','六','日'][day.weekday()] + ")"
        tasks_for_day = schedule_by_day[day]
        data_dict[day_str] = tasks_for_day + [''] * (max_rows - len(tasks_for_day))

    df = pd.DataFrame(data_dict)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='周工作计划')
        worksheet = writer.sheets['周工作计划']
        
        header_font = Font(bold=True)
        content_font = Font(bold=True)
        thin_side = Side(style='thin', color="000000")
        full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        alignment = Alignment(wrap_text=True, vertical='top')
        band_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        for col_idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 40
            
            header_cell = worksheet[f"{column_letter}1"]
            header_cell.font = header_font
            header_cell.border = full_border
            
            for row_idx in range(2, max_rows + 2):
                cell = worksheet[f"{column_letter}{row_idx}"]
                cell.alignment = alignment
                cell.border = full_border
                
                task_pair_index = (row_idx - 2) // 2
                if (row_idx - 2) % 2 == 0:
                    cell.font = content_font
                    if task_pair_index % 2 == 1:
                        cell.fill = band_fill
                else:
                    if task_pair_index % 2 == 1:
                        cell.fill = band_fill
    
    output.seek(0)
    log_activity('导出Excel', f"导出了 {start_of_week} 到 {end_of_week} 的工作计划。")
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=work_schedule_{start_of_week}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@main_bp.route('/logs')
@login_required
@admin_required
def activity_logs():
    page = request.args.get('page', 1, type=int)
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).paginate(page=page, per_page=20)
    return render_template('main/logs.html', logs=logs, title="操作日志")